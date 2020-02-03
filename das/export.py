from django.shortcuts import render, get_list_or_404
from django.views.decorators.http import require_POST
from django.http import HttpResponse, HttpResponseBadRequest
from django.utils import timezone

from rest_framework_jwt.serializers import VerifyJSONWebTokenSerializer #, ValidationError

import pytz
import datetime
import json
import openpyxl
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from . import models

def init_excel():
    return openpyxl.Workbook()

def parse_conf(req):
    data = json.loads(req.body.decode("utf-8"))

    tzinfo = pytz.timezone(data['timezone'])
    ts_from = int(data['ts_from']) # datetime.datetime.strptime(data['date_from'], "%Y-%m-%dT%H:%M:%S.%fZ")
    ts_to = int(data['ts_to']) # datetime.datetime.strptime(data['date_to'], "%Y-%m-%dT%H:%M:%S.%fZ")

    hide_null = 'hide_null' in data and data['hide_null']

    items = { int(id) for id in data['items'] }
    if not items:
        raise

    schemes = get_list_or_404(models.Scheme, groups__scheme_group_user__user_id=req.user.id, id__in=data['schemes'])
    return tzinfo, ts_from, ts_to, hide_null, items, schemes

def get_doc(scheme_id, items, tzinfo, ts_from, ts_to, hide_null):
    from collections import OrderedDict
    doc = {}

    from django.db import connection
    cursor = connection.cursor()

    cursor.execute("""
        SELECT s.id, s.name, it.id, it.title, l.timestamp_msecs, l.value FROM das_log_data l 
        LEFT JOIN das_device_item di ON l.item_id = di.id 
        LEFT JOIN das_device_item_type it ON di.type_id = it.id 
        LEFT JOIN das_device_item_group g ON di.group_id = g.id 
        LEFT JOIN das_section s ON g.section_id = s.id 
        WHERE l.scheme_id = {0} AND di.type_id IN ({1}) AND l.timestamp_msecs >= '{2}' AND l.timestamp_msecs <= '{3}'
        ORDER BY l.timestamp_msecs DESC;
    """.format(scheme_id, ",".join(map(str,items)), ts_from, ts_to))
    rawData = cursor.fetchall()
    #items = Log_Data.objects.filter(item__type_id__in=items, date__range=[from_str, to_str]).order_by('date')
    #for item in items:
    for row in rawData:
#        sct_id = item.item.group.section_id
        sct_id = row[0]
        if not sct_id in doc:
            doc[sct_id] = {
#                    'name': item.item.group.section.name,
                    'name': row[1],
                    'types': {}
                    }

            #item_type = item.item.type_id
            item_type = row[2]
        if not item_type in doc[sct_id]['types']:
            doc[sct_id]['types'][item_type] = {
                    #'name': item.item.type.title,
                    'name': row[3],
                    'dates': OrderedDict()
                    }

            value = None
        try:
#            value = float(item.value)
            value = float(row[5])
        except:
            value = 0.
        if value != 0.0 or not hide_null:
#            date_str = item.date.strftime('%d.%m.%Y')
#            date = item.date.date()
            dt = datetime.datetime.fromtimestamp(int(row[4]) / 1000, tzinfo)
            date = dt.date()
            if not date in doc[sct_id]['types'][item_type]['dates']:
                doc[sct_id]['types'][item_type]['dates'][date] = []
            doc[sct_id]['types'][item_type]['dates'][date].append({
                #'date': item.date,
                'date': dt,
                'value': value,
                })

    itype_dates = {}
    for sct_id in doc:
        doc_sct = doc[sct_id]
        for item_type in doc_sct['types']:
            if not item_type in itype_dates:
                itype_dates[item_type] = {}
            t = doc_sct['types'][item_type]
            for date in t['dates']:
                count = len(t['dates'][date])
                if not date in itype_dates[item_type] or itype_dates[item_type][date] < count:
                    itype_dates[item_type][date] = count

    for sct_id in doc:
        doc_sct = doc[sct_id]
        for item_type in doc_sct['types']:
            t = doc_sct['types'][item_type]
            for g_date in itype_dates[item_type]:
                if not g_date in t['dates']:
                    t['dates'][g_date] = []

            for date in t['dates']:
                d_len = itype_dates[item_type][date] - len(t['dates'][date])
                while d_len:
                    t['dates'][date].append({ 'date': None, 'value': '' })
                    d_len -= 1
                    
            t['dates'] = OrderedDict(sorted(t['dates'].items()))

    return doc

def generate_sheet(ws, doc):
    from collections import OrderedDict

    last_date = ''
    first_row = None

    def set_cell(row, column, value, alignment=None, bold=None, merge_row=None, merge_column=None):
        c = ws.cell(row=row, column=column, value=value)
        if alignment != None:
            c.alignment = openpyxl.styles.Alignment(horizontal=alignment, vertical='center')

        if bold != None:
            c.font = openpyxl.styles.Font(bold=bold)

        if merge_row != None or merge_column != None:
            if merge_row == None:
                merge_row = row
            if merge_column == None:
                merge_column = column
            ws.merge_cells(start_row=row,start_column=column,end_row=merge_row,end_column=merge_column)
        return c

    def sum_fmt(a, b, column):
        letter = get_column_letter(column)
        return '=SUM({2}{0}:{2}{1})'.format(a, b, letter)

    def check_last_row(row_n, col_n):
        if not first_row:
            return row_n

        row_count = row_n - first_row

        row_n -= 1
        # merge date
        ws.merge_cells(start_row=first_row,start_column=col_n,end_row=row_n,end_column=col_n)

        # merge Всего
        set_cell(first_row, col_n + 3, sum_fmt(first_row, row_n, col_n + 2), 'center', None, row_n)

        # merge Кол-во
        set_cell(first_row, col_n + 4, row_count, 'center', None, row_n)
        return row_n + 1

    def finalize_date():
        pass
    def finalize_group():
        pass

    columns = [
            (u"Дата", 15),
            (u"Время", 15),
            (u"Значение", 15),
            (u"Всего", 15),
            (u"Кол-во", 15),
            ]
    col_len = len(columns) - 1

    #        from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

    for col_num in range((col_len + 1) * len(doc)):
        cn = col_num % (col_len + 1) 
        set_cell(1, col_num + 1, columns[cn][0], 'center', True)
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[cn][1]

    ws.freeze_panes = ws['A2']
    col_num = 1

    for sct_id in doc:
        doc_sct = doc[sct_id]

        row_num = 2

        # Строка имени секции
        set_cell(row_num, col_num, doc_sct['name'], 'center', True, None, col_num+col_len)

        for item_type in doc_sct['types']:
            t = doc_sct['types'][item_type]

            date_idx = 0
            item_type_row = row_num
            row_num += 1
            set_cell(row_num, col_num, t['name'], 'center', True, None, col_num + 2)

            for date in t['dates']:
                items = t['dates'][date]
                empty_row_count = 0
                for item in items:
                    row_num += 1
                    if item['date']:
                        time_str = item['date'].strftime('%H:%M:%S')
                    else:
                        time_str = ''
                        empty_row_count += 1
                    set_cell(row_num, col_num + 1, time_str)
                    set_cell(row_num, col_num + 2, item['value'])
                item_len = len(items)
                first_row = row_num - item_len + 1
                set_cell(first_row, col_num, date, 'center', None, row_num)
                set_cell(first_row, col_num + 3, sum_fmt(first_row, row_num, col_num + 2), 'center', None, row_num) # merge Всего
                set_cell(first_row, col_num + 4, item_len - empty_row_count, 'center', None, row_num) # merge Кол-во
            value = 0 if item_type_row == row_num else sum_fmt(item_type_row + 2, row_num + 1, col_num + 2)
            set_cell(item_type_row + 1, col_num + 3, value, 'center')

            value = 0 if item_type_row == row_num else sum_fmt(item_type_row + 2, row_num + 1, col_num + 4)
            set_cell(item_type_row + 1, col_num + 4, value, 'center')

#            row_num = check_last_row(row_num, col_num)
            first_row = None
            last_date = ''
        first_row = None
        last_date = ''

        col_num += col_len + 1


@require_POST
def export_excel(req):
    auth = req.META.get('HTTP_AUTHORIZATION')
    valid_data = VerifyJSONWebTokenSerializer().validate({'token': auth[4:]})
    req.user = valid_data['user']

    #try:
    tzinfo, ts_from, ts_to, hide_null, items, schemes = parse_conf(req)
    wb = init_excel()
    empty_sheet = wb.active

    scheme_len = 0
    scheme_name = None
    for scheme in schemes:
        scheme_len += 1
        scheme_name = scheme.name
        doc = get_doc(scheme.id, items, tzinfo, ts_from, ts_to, hide_null)

        ws = wb.create_sheet(scheme.title[:31])
        generate_sheet(ws, doc)

    get_ts_str = lambda ts: datetime.datetime.fromtimestamp(ts/1000).strftime('%d.%m.%Y_%H.%M')
    file_name = (scheme_name + '__') if scheme_len == 1 else ''
    file_name += get_ts_str(ts_from) + '__' + get_ts_str(ts_to)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="{0}.xlsx"'.format(file_name)
    wb.remove_sheet(empty_sheet)
    wb.save(response)
    return response
    #except:
    #    return HttpResponseBadRequest()

